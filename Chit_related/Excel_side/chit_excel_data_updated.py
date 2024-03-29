import openpyxl
import os
from tkinter import filedialog as fd
import re
from openpyxl.styles import Font

def getFilePath():
    # selecting the file using the askopenfilename() method of filedialog
    the_file = fd.askopenfilename(
        title="Select an excel file",
        filetypes=[("Excel files", "*.xlsx")]
    )
    # getting path of a file using the startfile() method of the os module
    file_path = os.path.abspath(the_file)
    return file_path
    # os.startfile(os.path.abspath(the_file))


def add_suffix_to_filename(file_path, suffix):
    directory_path = os.path.dirname(file_path)
    filename, extension = os.path.splitext(os.path.basename(file_path))
    updated_filename = filename + suffix + extension
    updated_file_path = os.path.join(directory_path, updated_filename)
    return updated_file_path

in_file_path = fr'{getFilePath()}'
print(in_file_path)
workbook = openpyxl.load_workbook(in_file_path)


def create_sheet(sheet_num: int):
    work_Sheet = workbook[f'Sheet{str(sheet_num)}']
    return work_Sheet



def getNumData(start_row: int, column: str, workSheet):
    start_cell, end_row, data_num = workSheet[column + str(start_row)], start_row, {}
    itr_cell = start_cell
    while itr_cell.value is not None:
        data_num[str(end_row - start_row + 1)] = itr_cell.value
        end_row += 1
        itr_cell = workSheet[column + str(end_row)]
    return data_num
    # length = (end_row - start_row) + 1


data2 = getNumData(3, 'B', create_sheet(1))
num_list = [str(num) for num in range(1, len(data2) + 1)]
totals = []


def autoFillSum(start_row: int, end_row: int, column: str, workSheet):
    total_sum = 0
    for row in range(start_row, end_row + 1):
        cell = workSheet[column + str(row)]
        value = cell.value
        if value is not None:
            str_list = re.findall(r'\d+', value)
            temp_sum, count = 0, 0
            for digit in str_list:
                if digit != '' and digit in num_list:
                    count += 1
                    temp_sum += data2[digit]
                else:
                    temp_sum = 0
                    break
            total_sum += temp_sum
            next_cell = workSheet[chr(ord(column) + 1) + str(row)]
            next_cell.value = temp_sum
    sum_cell = workSheet[chr(ord(column) + 1) + str(end_row + 1)]
    sum_cell.font, sum_cell.value = Font(bold=True), total_sum
    totals.append(total_sum)


# print(getNumData(3, 'B', worksheet1))
autoFillSum(3, 32, 'B', create_sheet(2))
autoFillSum(3, 32, 'E', create_sheet(2))
autoFillSum(3, 40, 'B', create_sheet(3))
autoFillSum(3, 40, 'E', create_sheet(3))
# autoFillSum(5, 128, 'B', worksheet4)


# Grand total
create_sheet(3)['F43'].font = Font(bold=True, italic=True, size=14)
create_sheet(3)['E43'].value, create_sheet(3)['F43'].value = "GRAND TOTAL", sum(totals)

#save file path
# out_file_path = fr'{os.path.dirname(in_file_path)}\updated_chit.xlsx'
out_file_path = add_suffix_to_filename(in_file_path, "_updated")
# save file

workbook.save(out_file_path)


