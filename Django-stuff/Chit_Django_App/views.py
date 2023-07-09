import openpyxl
import re
from openpyxl.styles import Font
from django.shortcuts import render
from django.http import HttpResponse

def index(request):
    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        workbook = openpyxl.load_workbook(myfile)

        def create_sheet(sheet_num: int):
            work_Sheet = workbook[f'Sheet{str(sheet_num)}']
            return work_Sheet

        def getNumData(start_row: int, column: str, workSheet):
            start_cell, end_row, data_num = workSheet[column + str(start_row)], start_row, {}
            itr_cell = start_cell
            while itr_cell.value is not None:
                data_num[str(end_row - start_row + 1)] = itr_cell.value
                end_row += 1
                itr_cell = workSheet[column + str(end_row)]
            return data_num

        data2 = getNumData(3, 'B', create_sheet(1))
        num_list = [str(num) for num in range(1, len(data2) + 1)]
        totals = []

        def autoFillSum(start_row: int, end_row: int, column: str, workSheet):
            total_sum = 0
            for row in range(start_row, end_row + 1):
                cell = workSheet[column + str(row)]
                value = cell.value
                if value is not None:
                    str_list = re.findall(r'\d+', value)
                    temp_sum, count = 0, 0
                    for digit in str_list:
                        if digit != '' and digit in num_list:
                            count += 1
                            temp_sum += data2[digit]
                        else:
                            temp_sum = 0
                            break
                    total_sum += temp_sum
                    next_cell = workSheet[chr(ord(column) + 1) + str(row)]
                    next_cell.value = temp_sum
            sum_cell = workSheet[chr(ord(column) + 1) + str(end_row + 1)]
            sum_cell.font, sum_cell.value = Font(bold=True), total_sum
            totals.append(total_sum)

        autoFillSum(3, 32, 'B', create_sheet(2))
        autoFillSum(3, 32, 'E', create_sheet(2))
        autoFillSum(3, 40, 'B', create_sheet(3))
        autoFillSum(3, 40, 'E', create_sheet(3))

        create_sheet(3)['F43'].font = Font(bold=True, italic=True, size=14)
        create_sheet(3)['E43'].value, create_sheet(3)['F43'].value = "GRAND TOTAL", sum(totals)

        workbook.save("chit_data_updated.xlsx")

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=chit_data_updated.xlsx'
        workbook.save(response)
        return response

    return render(request, 'index.html')
